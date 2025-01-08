import os
import re

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import ImageFont
import tkinter as tk
from tkinter import filedialog, messagebox


class data:
    def __init__(self, danghao, wenjianbianhao, zerenzhe, wenjiantiming, riqi, yehao ,yeshu):
        self.danghao = danghao
        self.wenjianbianhao = wenjianbianhao
        self.zerenzhe = zerenzhe
        self.wenjiantiming = wenjiantiming
        self.riqi = riqi
        self.yehao = yehao
        self.yeshu = yeshu


def clearTab(file_name: str):
    # 获取当前目录
    current_directory = os.getcwd()
    # 构建文件路径
    file_path = os.path.join(current_directory, file_name)
    # 检查文件是否存在并删除
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_name} 已删除")
        pass
    else:
        print(f"{file_name} 不存在")

    # newTab = Workbook()
    # sheet = newTab.active
    # sheet.title = "Sheet1"
    # newTab.save(file_name)
    pass


def getForm():
    workbook = load_workbook('tmp.xlsx')
    sheet = workbook.active

    # 打印1-14行的行高
    for row in range(1, 15):
        row_height = sheet.row_dimensions[row].height
        print(f"Row {row} height: {row_height}")

    # 打印A-G列的列宽
    for col in 'ABCDEFG':
        col_width = sheet.column_dimensions[col].width
        print(f"Column {col} width: {col_width}")


def calculate_column_width_pil(font_name: str, font_size: int, byte_width: float) -> float:
    """
    使用 PIL 根据字体和字号计算 Excel 列宽。

    :param font_name: 字体名称，例如 "宋体"
    :param font_size: 字号大小，例如 11
    :param byte_width: Excel 中的列宽字节大小，例如 4.19
    :return: 适合的列宽值
    """
    # 加载字体
    try:
        font = ImageFont.truetype(font_name, font_size)
    except OSError:
        raise ValueError(f"无法加载字体 '{font_name}'，请确保字体文件存在并路径正确。")

    # 测量一个字符的宽度（以 '0' 为基准，因为 Excel 列宽基于 '0' 的宽度）
    char_width = font.getbbox("0")[2] - font.getbbox("0")[0]

    # 微软默认的列宽基准：Calibri 字体，11 号，'0' 的宽度约为 7 像素，对应列宽 1
    default_char_width = 7  # Calibri 11 号字体中 '0' 的宽度
    default_column_width = 1  # 对应的列宽

    # 计算当前字体的列宽比例
    width_ratio = char_width / default_char_width

    # 根据目标字节宽度计算最终列宽
    adjusted_width = byte_width * width_ratio
    return adjusted_width


def fitData(mainTab=None, sheetName=None, data_list=None):
    if data_list is None or mainTab is None:
        return
    else:
        mysheet = mainTab[sheetName]

        mysheet["A1"] = "卷内目录"
        mysheet["A2"] = "      档号 " + data_list[0].danghao
        mysheet["A3"] = "序\n号"
        mysheet["B3"] = "文件\n编号"
        mysheet["C3"] = "责任者"
        mysheet["D3"] = "文件题名"
        mysheet["E3"] = "日期"
        mysheet["F3"] = "页号"
        mysheet["G3"] = "备注"

        data_listALLPage = 0
        for page_num in data_list:
            data_listALLPage += page_num.yeshu
        print(data_listALLPage)


        if len(data_list) < 11:

            first = 4
            for row in range(first, first + len(data_list)):
                data_list_index = row - first
                if data_list_index == len(data_list) -1:
                    mysheet["F" + str(row)].value = str(data_list[data_list_index].yehao) + "/" + str(data_listALLPage)
                else:
                    mysheet["F" + str(row)].value = data_list[data_list_index].yehao

                mysheet["A" + str(row)].value = data_list_index + 1
                mysheet["C" + str(row)].value = data_list[data_list_index].zerenzhe
                mysheet["D" + str(row)].value = data_list[data_list_index].wenjiantiming
                mysheet["E" + str(row)].value = data_list[data_list_index].riqi
                mysheet["B" + str(row)].value = data_list[data_list_index].wenjianbianhao
        else:
            the_lenth = len(data_list)
            all_pages = int((the_lenth + 9) / 10)

            for now_page in range(1, all_pages + 1):
                tabindex = (now_page - 1) * 11 + 4
                listindex = (now_page - 1) * 10

                tablast = (now_page) * 11 + 4 - 1
                mysheet["A" + str(tablast)].value = "第" + str(now_page) + "页/共" + str(all_pages) + "页"

                for row in range(tabindex, tabindex + 10):

                    if listindex >= len(data_list):
                        break
                    else:
                        if listindex == len(data_list) -1:
                            mysheet["F" + str(row)].value = str(data_list[listindex].yehao) + "/" + str(data_listALLPage)
                        else:
                            mysheet["F" + str(row)].value = data_list[listindex].yehao

                        mysheet["A" + str(row)].value = listindex + 1
                        mysheet["C" + str(row)].value = data_list[listindex].zerenzhe
                        mysheet["D" + str(row)].value = data_list[listindex].wenjiantiming
                        mysheet["E" + str(row)].value = data_list[listindex].riqi
                        mysheet["B" + str(row)].value = data_list[listindex].wenjianbianhao
                        listindex += 1

    return mainTab


def decorate(mainTab: Workbook, sheetName=None, data_list=None):
    if len(data_list) is None:
        return
    mainsheet = mainTab[sheetName]
    # 设置行高
    mainsheet.row_dimensions[1].height = 45.0
    mainsheet.row_dimensions[2].height = 16.5
    mainsheet.row_dimensions[3].height = 56.25
    mainsheet.row_dimensions[4].height = 72.0
    mainsheet.row_dimensions[5].height = 72.0
    mainsheet.row_dimensions[6].height = 72.0
    mainsheet.row_dimensions[7].height = 72.0
    mainsheet.row_dimensions[8].height = 72.0
    mainsheet.row_dimensions[9].height = 72.0
    mainsheet.row_dimensions[10].height = 72.0
    mainsheet.row_dimensions[11].height = 72.0
    mainsheet.row_dimensions[12].height = 72.0
    mainsheet.row_dimensions[13].height = 72.0
    mainsheet.row_dimensions[14].height = 20
    # 设置列宽
    mainsheet.column_dimensions['A'].width = 5.89814814814815
    mainsheet.column_dimensions['B'].width = 11.6944444444444
    mainsheet.column_dimensions['C'].width = 10.6203703703704
    mainsheet.column_dimensions['D'].width = 37.5092592592593
    mainsheet.column_dimensions['E'].width = 11.6944444444444
    mainsheet.column_dimensions['F'].width = 11.6944444444444
    mainsheet.column_dimensions['G'].width = 5.89814814814815

    # 设置单元格格式和样式

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # A1
    mainsheet["A1"].font = Font(name="宋体", bold=True, size=20)
    mainsheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # A2
    mainsheet["A2"].font = Font(name="宋体", bold=True, size=12)
    mainsheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # A3, B3, G3
    for cell in ["A3", "B3", "G3"]:
        mainsheet[cell].font = Font(name="宋体", bold=True, size=12)
        mainsheet[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # C3, D3, E3, F3
    for cell in ["C3", "D3", "E3", "F3"]:
        mainsheet[cell].font = Font(name="宋体", bold=True, size=12)
        mainsheet[cell].alignment = Alignment(horizontal="center", vertical="center")

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = mainsheet[f"{col}3"]
        cell.border = thin_border

    if len(data_list) <= 10:
        # A4-A13, B4-B13, C4-C13
        for row in range(4, 14):
            for col in ["A", "B", "C"]:
                mainsheet[f"{col}{row}"].font = Font(name="宋体", size=10)
                mainsheet[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # D4-D13
        for row in range(4, 14):
            mainsheet[f"D{row}"].font = Font(name="宋体", size=10)
            mainsheet[f"D{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # E4-E13, F4-F13, G4-G13
        for row in range(4, 14):
            for col in ["E", "F", "G"]:
                mainsheet[f"{col}{row}"].font = Font(name="宋体", size=10)
                mainsheet[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        for row in range(4, 14):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                cell = mainsheet[f"{col}{row}"]
                cell.border = thin_border
        pass
    else:

        the_lenth = len(data_list)
        all_pages = int((the_lenth + 9) / 10)

        for now_page in range(1, all_pages + 1):
            tabindex = (now_page - 1) * 11 + 4
            tablast = (now_page) * 11 + 4 - 1

            for row in range(tabindex, tabindex + 10):
                mainsheet.row_dimensions[row].height = 72.0

                mainsheet["A" + str(row)].font = Font(name="宋体", size=10)
                mainsheet["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                mainsheet["C" + str(row)].font = Font(name="宋体", size=10)
                mainsheet["C" + str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                mainsheet["B" + str(row)].font = Font(name="宋体", size=10)
                mainsheet["B" + str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                mainsheet["D" + str(row)].font = Font(name="宋体", size=10)
                mainsheet["D" + str(row)].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                mainsheet["E" + str(row)].font = Font(name="宋体", size=10)
                mainsheet["E" + str(row)].alignment = Alignment(horizontal="center", vertical="center")

                mainsheet["F" + str(row)].font = Font(name="宋体", size=10)
                mainsheet["F" + str(row)].alignment = Alignment(horizontal="center", vertical="center")

                mainsheet["G" + str(row)].font = Font(name="宋体", size=10)
                mainsheet["G" + str(row)].alignment = Alignment(horizontal="center", vertical="center")

            for row in range(tabindex, tablast):
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    cell = mainsheet[f"{col}{row}"]
                    cell.border = thin_border

            mainsheet.merge_cells(f'A{tablast}:G{tablast}')
            mainsheet[f'A{tablast}'].alignment = Alignment(horizontal="center", vertical="center")

        pass

    mainsheet.merge_cells('A1:G1')
    mainsheet.merge_cells('A2:G2')

    cm_to_inch = 1 / 2.54
    margins = PageMargins(
        left=2.5 * cm_to_inch, right=1 * cm_to_inch,
        top=1.1 * cm_to_inch, bottom=1.9 * cm_to_inch,
        header=0 * cm_to_inch, footer=0.8 * cm_to_inch
    )
    mainsheet.page_margins = margins
    mainsheet.print_title_rows = '1:3'
    mainsheet.page_setup.scale = 93
    mainsheet.page_setup.paperSize = mainsheet.PAPERSIZE_A4

    return mainTab


class row_insource:

    def __init__(self,
                 quanzonghao,
                 xiangmudaihao,
                 danghao,
                 fenleihao,
                 xiangmumingcheng,
                 anjuanhao,
                 taoshu,
                 juanneishunxuhao,
                 wenjiantiming,
                 wenjianbianhao,
                 zerenzhe,
                 riqi,
                 yehao,
                 yeshu,
                 lijuandanwei,
                 baoguanriqi,
                 miji,
                 beizhu):
        self.quanzonghao = quanzonghao
        self.xiangmudaihao = xiangmudaihao
        self.danghao = danghao
        self.fenleihao = fenleihao
        self.xiangmumingcheng = xiangmumingcheng
        self.anjuanhao = anjuanhao
        self.taoshu = taoshu
        self.juanneishunxuhao = juanneishunxuhao
        self.wenjiantiming = wenjiantiming
        self.wenjianbianhao = wenjianbianhao
        self.zerenzhe = zerenzhe
        self.riqi = riqi
        self.yehao = yehao
        self.yeshu = yeshu
        self.lijuandanwei = lijuandanwei
        self.baoguanriqi = baoguanriqi
        self.miji = miji
        self.beizhu = beizhu


def sort_key(danghao):
    match = re.search(r'(\d+)$', danghao)
    return int(match.group(1)) if match else float('inf')


def doRead(file_path):
    mybook = load_workbook(file_path, data_only=True)
    mysheet = mybook["Sheet1"]

    data_list = []
    for row in mysheet.iter_rows(min_row=2, values_only=True):
        row = [' ' if cell is None else cell for cell in row]
        data_item = row_insource(*row)
        data_list.append(data_item)

    return data_list


def transForm(sum_sublist):
    data_list = []

    for item in sum_sublist:
        dataitem = data(danghao=item.danghao,
                        wenjianbianhao=item.wenjianbianhao,
                        zerenzhe=item.zerenzhe,
                        wenjiantiming=item.wenjiantiming,
                        riqi=item.riqi,
                        yehao=item.yehao,
                        yeshu=item.yeshu)
        data_list.append(dataitem)
    return data_list


def domain(source_name, output_name):
    if source_name is None:
        return print("file_path is None")

    data_list = doRead(source_name)

    danghao_set = set()
    # 遍历 the_complete 列表，提取不同的档号
    for item in data_list:
        danghao_set.add(item.danghao)
    # 将集合转换为列表
    danghao_list = list(danghao_set)
    sorted_danghao_list = sorted(danghao_list, key=sort_key)

    # 最终需要存储的文件
    theSumBook = Workbook()


    index = 1

    for item in sorted_danghao_list:
        # 找到同档号
        sum_sublist = []
        for sub_item in data_list:
            if (sub_item.danghao == item):
                sum_sublist.append(sub_item)

        # 转换为data类型
        data_item = transForm(sum_sublist)

        # 当数据为空时直接跳过。剩下的所有操作，进入下一个循环
        if data_item is None:
            continue

        # 当第一次写入时，要写入 theSumBook 中的 Sheet 卷内目录1
        if len(theSumBook.sheetnames) == 1 and theSumBook.active.title == "Sheet":
            mysheet = theSumBook.active
            mysheet.title = "卷内目录" + str(index)
        else:
            # 当第N次写入时，要写入 theSumBook 中的 Sheet 卷内目录N
            mysheet = theSumBook.create_sheet(title="卷内目录" + str(index))


        ## 设置默认行高和列宽
        #default_row_height = 15.6
        #default_column_width = 9.23484848484848
        ## 设置所有行的默认行高
        #for row in range(1, mysheet.max_row + 1):
        #    mysheet.row_dimensions[row].height = default_row_height
        ## 设置所有列的默认列宽
        #for col in range(1, mysheet.max_column + 1):
        #    mysheet.column_dimensions[get_column_letter(col)].width = default_column_width


        theSumBook = fitData(data_list=data_item, mainTab=theSumBook, sheetName=mysheet.title)
        theSumBook = decorate(data_list=data_item, mainTab=theSumBook, sheetName=mysheet.title)
        all_pages = (mysheet.max_row - 3) / 11
        mysheet.max_row

        row_groups = []
        row_in = 14
        while (row_in <= mysheet.max_row):
            row_groups.append((row_in - 10, row_in))
            row_in += 11

        for start_row, end_row in row_groups:
            print_area = f'A{start_row}:G{end_row - 1}'  # Adjust as needed for your columns
            mysheet.print_area += f',{print_area}' if mysheet.print_area else print_area

        mysheet.page_setup.fitToWidth = 1  # 将宽度调整为1页
        mysheet.page_setup.fitToHeight = False  # 高度限制页数

        index += 1
    theSumBook.save(output_name)
    pass


def select_source_file():
    file_path = filedialog.askopenfilename(
        title="选择 案卷目录.xlsx 或源文件",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        source_entry.delete(0, tk.END)
        source_entry.insert(0, file_path)

def generate_file():
    source_name = source_entry.get()
    output_name = output_entry.get()
    if not source_name:
        messagebox.showwarning("警告", "请先选择 案卷目录.xlsx 文件")
        return
    if not output_name:
        output_name = "卷内总目录生成卷内目录.xlsx"
    # 这里调用你的逻辑函数
    clearTab(output_name)
    domain(source_name, output_name)
    messagebox.showinfo("成功", f"文件已生成：{output_name}")

if __name__ == "__main__":
    root = tk.Tk()
    root.title("生成 卷内总目录生成卷内目录.xlsx")

    # 创建一个框架来组织控件
    frame = tk.Frame(root)
    frame.grid(padx=20, pady=20)

    # 第一行：选择source文件
    tk.Label(frame, text="选择 案卷目录.xlsx 文件:").grid(row=0, column=0, sticky="w", padx=10, pady=10)
    source_entry = tk.Entry(frame, width=50)
    source_entry.grid(row=0, column=1, padx=10, pady=10)
    tk.Button(frame, text="浏览", command=select_source_file).grid(row=0, column=2, padx=10, pady=10)

    # 第二行：警告信息
    tk.Label(frame, text="同名文件会被新文件覆盖，请做好备份！！！！", fg="red").grid(row=1, column=0, columnspan=3, sticky="w", padx=10, pady=10)

    # 第三行：输出文件名
    tk.Label(frame, text="输出文件名 卷内总目录生成卷内目录:").grid(row=2, column=0, sticky="w", padx=10, pady=10)
    output_entry = tk.Entry(frame, width=50)
    output_entry.insert(0, "卷内总目录生成卷内目录.xlsx")
    output_entry.grid(row=2, column=1, padx=10, pady=10)

    # 第四行：下一步按钮
    tk.Button(frame, text="下一步", command=generate_file).grid(row=3, column=1, padx=10, pady=20)

    root.mainloop()