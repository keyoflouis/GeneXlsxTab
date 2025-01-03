import os
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Border, Font, Alignment
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.page import PageMargins


class dataTab2:
    def __init__(self, danghao, jianshu, yeshu):
        self.danghao = danghao
        self.jianshu = jianshu
        self.yeshu = yeshu


def getform():
    # 打开Excel文件
    workbook = openpyxl.load_workbook('tab2.xlsx')
    sheet = workbook.active

    # 打印第一列的行高和列宽
    for row in range(1, 28):
        cell = sheet.cell(row=row, column=1)
        row_height = sheet.row_dimensions[row].height
        col_width = sheet.column_dimensions[cell.column_letter].width
        print(f"Row {row}: Height = {row_height}, Column A: Width = {col_width}")


def creatTable(tab2Name: str):
    tab2 = Workbook()
    activesheet = tab2.active
    activesheet.title = "Sheet1"
    tab2.save(tab2Name)


def oganizeDataTab2(data: dataTab2):
    newGrid = Workbook()
    activeSheet = newGrid.active

    activeSheet['A1'].value = "卷内备考表"
    activeSheet['A4'].value = "互见号:"
    activeSheet['A6'].value = "说明:"
    activeSheet['A23'].value = "                                      立卷人："
    activeSheet['A24'].value = "                                                   年    月    日"
    activeSheet['A25'].value = "                                      检查人："
    activeSheet['A26'].value = "                                                   年    月    日"

    activeSheet['A2'].value = "档号:" + str(data.danghao)

    # 需要在data.yeshu加上下划线
    # activeSheet['A7'].value = "      本卷档共有文件" + str(data.jianshu) + "件，共" + str(data.yeshu) + "页。"

    # 创建富文本格式
    normal_text = "      本卷档共有文件"
    jianshu_text = "  " + str(data.jianshu) + "  "
    # 带下划线的部分
    mid_text = "件，共"
    yeshu_text = "  " + str(data.yeshu) + "  "
    last_text = "页。"

    songti_13 = InlineFont(rFont="宋体", sz=13)

    # 创建带下划线的字体
    underline_font = InlineFont(u='single', rFont='宋体', sz=13)  # 使用单下划线
    # 创建富文本对象
    rich_text = CellRichText(
        TextBlock(songti_13, normal_text),
        TextBlock(underline_font, jianshu_text),
        TextBlock(songti_13, mid_text),
        TextBlock(underline_font, yeshu_text),
        TextBlock(songti_13, last_text)
    )

    # 将富文本赋值给单元格
    activeSheet['A7'] = rich_text
    # newGrid.save("test2.xlsx")
    return newGrid


def copyDatatoTab2(mainTab: Workbook, subTab: Workbook, startIndex: str):
    mainsheet = mainTab.active
    subsheet = subTab.active

    index = column_index_from_string(startIndex)
    A = get_column_letter(index)

    mainsheet[A + '1'].value = subsheet['A1'].value
    mainsheet[A + '4'].value = subsheet['A4'].value
    mainsheet[A + '6'].value = subsheet['A6'].value
    mainsheet[A + '23'].value = subsheet['A23'].value
    mainsheet[A + '24'].value = subsheet['A24'].value
    mainsheet[A + '25'].value = subsheet['A25'].value
    mainsheet[A + '26'].value = subsheet['A26'].value

    mainsheet[A + '2'].value = subsheet['A2'].value
    mainsheet[A + '7'].value = subsheet['A7'].value

    return mainTab


def decorateTab2(mainTab: Workbook, startIndex: str):
    mainsheet = mainTab.active

    # 获取起始列的字母
    index = column_index_from_string(startIndex)
    A = get_column_letter(index)

    # 设置行高
    for row in range(1, 28):
        mainsheet.row_dimensions[row].height = 30.0

    # 设置列宽
    mainsheet.column_dimensions[A].width = 82.5666666666667

    # 定义边框样式
    thin = Side(border_style="thin", color="000000")
    border_top = Border(top=thin)
    border_bottom = Border(bottom=thin)
    border_right = Border(right=thin)

    # 设置字体和对齐方式
    font_a1 = Font(name="宋体", size=22, bold=True)
    font_a2 = Font(name="宋体", size=14, bold=True)
    font_a3_a5_a8_a22_a27 = Font(name="宋体", size=12)
    font_a4_a6_a7_a23_a26 = Font(name="宋体", size=13)

    align_center = Alignment(vertical="center", horizontal="center")
    align_left = Alignment(vertical="center", horizontal="left")

    # 设置 A1 单元格样式
    mainsheet[f"{A}1"].font = font_a1
    mainsheet[f"{A}1"].alignment = align_center

    # 设置 A2 单元格样式
    mainsheet[f"{A}2"].font = font_a2
    mainsheet[f"{A}2"].alignment = align_left

    # 设置 A3 到 A27 的样式
    for row in range(3, 28):
        cell = mainsheet[f"{A}{row}"]

        # 设置字体
        if row in [3, 5] or 8 <= row <= 22 or row == 27:
            cell.font = font_a3_a5_a8_a22_a27
            cell.alignment = align_center
        elif row in [4, 6, 7] or 23 <= row <= 26:
            cell.font = font_a4_a6_a7_a23_a26
            cell.alignment = align_left

        # 设置边框
        if row == 3:
            cell.border = Border(top=thin, right=thin,left=thin)
        elif row == 27:
            cell.border = Border(bottom=thin, right=thin,left=thin)
        else:
            cell.border = Border(right=thin,left=thin)

    return mainTab


def doWriteToTab2(data: dataTab2, filepath: str, start: str):
    mybook = load_workbook(filepath, rich_text=True)

    sheet = mybook["Sheet1"]
    sheet.page_margins = PageMargins(left=1.78 / 2.54,  # 转换为英寸
                                    right=1.78 / 2.54,
                                    top=1.91 / 2.54,
                                    bottom=1.91 / 2.54,
                                    header=0.76 / 2.54,
                                    footer=0.76 / 2.54)

    temptab = oganizeDataTab2(data)
    mybook = copyDatatoTab2(mybook, temptab, start)
    mybook = decorateTab2(mybook, start)

   # # 获取最大列数
   # max_col = sheet.max_column

   # # 设置打印区域
   # sheet.print_area = f'A1:{get_column_letter(max_col)}27'

   # # 设置每列单独打印
   # for col in range(1, 28):
   #     sheet.page_setup.fitToWidth = 1
   #     sheet.page_setup.fitToHeight = False
   #     sheet.page_setup.orientation = 'portrait'


    mybook.save(filepath)

    return get_column_letter(column_index_from_string(start) + 1)


if __name__ == "__main__":

    testdata = dataTab2("GYHSD-JL2020-JL-1", 2, 53)

    testName = "test2.xlsx"
    current_directory = os.getcwd()
    file_path = os.path.join(current_directory, testName)

    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{testName} 已删除")
    else:
        print(f"{testName} 不存在")

    creatTable(testName)
    doWriteToTab2(testdata, testName, "A")
