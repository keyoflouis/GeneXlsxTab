import os
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from WriteToTab1 import dataTab1
from WriteToTab1 import doWriteToTab1

from WriteToTab2 import dataTab2
from WriteToTab2 import doWriteToTab2
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

def format_date(date_str):
    return datetime.strptime(date_str, "%Y%m%d").strftime("%Y年%m月%d日")


def getDataFromSource_ForTab1(source_file):
    source = load_workbook(source_file)
    source_tab = source["Sheet1"]
    datas = []

    # 获取表头
    header = {cell.value: cell.column_letter for cell in source_tab[1]}  # 假设第一行为表头

    for row in range(2, source_tab.max_row + 1):  # 从第二行开始读取
        danghao = source_tab[f"{header['档号']}{row}"].value
        anjuantiming = source_tab[f"{header['案卷题名']}{row}"].value
        baoguanqixian = source_tab[f"{header['保管期限']}{row}"].value

        start_time = source_tab[f"{header['起始日期']}{row}"].value
        end_time = source_tab[f"{header['终止日期']}{row}"].value
        if start_time is None or end_time is None:
            qizhiriqi = "—"
        else:
            qizhiriqi = format_date(str(start_time)) + "至" + format_date(str(end_time))

        lijuandanwei = source_tab[f"{header['立卷单位']}{row}"].value
        miji = source_tab[f"{header['密级']}{row}"].value

        item = dataTab1(DangHao=danghao,
                        AnJuanTiMing=anjuantiming,
                        LiJuanDanWei=lijuandanwei,
                        QiZhiRiQi=qizhiriqi,
                        BaoGuanQiXian=baoguanqixian,
                        MiJi=miji)

        datas.append(item)
    return datas


def getDataFromSource_ForTab2(source_file):
    source = load_workbook(source_file)
    source_tab = source["Sheet1"]
    datas = []

    # 获取表头
    header = {cell.value: cell.column_letter for cell in source_tab[1]}  # 假设第一行为表头

    for row in range(2, source_tab.max_row + 1):  # 从第二行开始读取
        danghao = source_tab[f"{header['档号']}{row}"].value

        zongjianshu = source_tab[f"{header['总件数']}{row}"].value
        zongyeshu = source_tab[f"{header['总页数']}{row}"].value

        if zongjianshu is None or zongyeshu is None:
            continue
        else:
            item = dataTab2(danghao=danghao, jianshu=zongjianshu, yeshu=zongyeshu)
            datas.append(item)

    return datas


def creatXlsx(file_name):
    # 获取当前目录
    current_directory = os.getcwd()
    # 构建文件路径
    file_path = os.path.join(current_directory, file_name)
    # 检查文件是否存在并删除
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_name} 已删除")
        pass
    else:
        print(f"{file_name} 不存在")

    newTab = Workbook()
    sheet = newTab.active
    sheet.title = "Sheet1"
    newTab.save(file_name)
    pass


def doMain(source_file, gene_file_Tab1, gene_file_Tab2):
    datas_tab1 = getDataFromSource_ForTab1(source_file)
    datas_tab2 = getDataFromSource_ForTab2(source_file)

    startIndex = 'A'

    for item in datas_tab1:
        startIndex = doWriteToTab1(item, gene_file_Tab1, startIndex)

    startIndex = 'A'

    for item in datas_tab2:
        startIndex = doWriteToTab2(item, gene_file_Tab2, startIndex)


#if __name__ == "__main__":
#    source_file = "source.xlsx"
#    gene_name_Tab1 = "test.xlsx"
#    gene_name_Tab2 = "test2.xlsx"
#
#    creatXlsx(gene_name_Tab1)
#    creatXlsx(gene_name_Tab2)
#
#    doMain(source_file, gene_name_Tab1, gene_name_Tab2)

def select_source_file():
    file_path = filedialog.askopenfilename(
        title="选择 案卷目录.xlsx 或源文件",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        source_entry.delete(0, tk.END)
        source_entry.insert(0, file_path)

def generate_files():
    source_file = source_entry.get()
    gene_name_Tab1 = tab1_entry.get()
    gene_name_Tab2 = tab2_entry.get()

    if not source_file:
        messagebox.showwarning("警告", "请选择 案卷目录.xlsx 文件")
        return

    if not gene_name_Tab1:
        gene_name_Tab1 = "test.xlsx"
    if not gene_name_Tab2:
        gene_name_Tab2 = "test2.xlsx"

    # 这里调用你的逻辑函数
    creatXlsx(gene_name_Tab1)
    creatXlsx(gene_name_Tab2)
    doMain(source_file, gene_name_Tab1, gene_name_Tab2)

    messagebox.showinfo("完成", "文件生成成功")


source_entry:None
tab1_entry:None
tab2_entry:None
if __name__ == "__main__":
    root = tk.Tk()
    root.title("生成 卷案生成.xlsx 和 卷内生成备考表.xlsx")

    tk.Label(root, text="选择 案卷目录.xlsx 文件:").grid(row=0, column=0, padx=10, pady=10)
    source_entry = tk.Entry(root, width=50)
    source_entry.grid(row=0, column=1, padx=10, pady=10)
    tk.Button(root, text="浏览", command=select_source_file).grid(row=0, column=2, padx=10, pady=10)

    tk.Label(root, text="同名文件会被新文件覆盖，请做好备份！！", fg="red").grid(row=1, column=0, columnspan=3, sticky="w", padx=10, pady=10)

    tk.Label(root, text="生成文件名 卷案生成:").grid(row=2, column=0, padx=10, pady=10)
    tab1_entry = tk.Entry(root, width=50)
    tab1_entry.insert(0, "卷案生成.xlsx")
    tab1_entry.grid(row=2, column=1, padx=10, pady=10)

    tk.Label(root, text="生成文件名 卷内生成备考表:").grid(row=3, column=0, padx=10, pady=10)
    tab2_entry = tk.Entry(root, width=50)
    tab2_entry.insert(0, "卷内生成备考表.xlsx")
    tab2_entry.grid(row=3, column=1, padx=10, pady=10)

    tk.Button(root, text="下一步", command=generate_files).grid(row=4, column=1, padx=10, pady=10)

    root.mainloop()



