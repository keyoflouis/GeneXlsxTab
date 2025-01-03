

from openpyxl import Workbook ,load_workbook
from openpyxl.styles import Alignment, Border, Side ,Font
from openpyxl.utils import get_column_letter ,column_index_from_string
from openpyxl.worksheet.page import PageMargins


class dataTab1:
    def __init__(self,DangHao,AnJuanTiMing,LiJuanDanWei,QiZhiRiQi,BaoGuanQiXian,MiJi):
        self.DangHao = DangHao
        self.AnJuanTiMing = AnJuanTiMing
        self.LiJuanDanWei = LiJuanDanWei
        self.QiZhiRiQi = QiZhiRiQi
        self.BaoGuanQiXian = BaoGuanQiXian
        self.MiJi = MiJi

def oganizeDataTab1(DangHao, AnJuanTiMing, LiJuanDanWei, QiZhiRiQi, BaoGuanQiXian, MiJi):

    uniteBook = Workbook()
    nowSheet = uniteBook.active

    nowSheet['B2'] = '档号：'
    nowSheet['B7'] ='立卷单位'
    nowSheet['B8'] ='起止日期'
    nowSheet['B9'] ='保管期限'
    nowSheet['B10'] ='密级'

    nowSheet['C2'] =DangHao
    nowSheet['B6'] =AnJuanTiMing
    nowSheet['C7'] =LiJuanDanWei
    nowSheet['C8'] =QiZhiRiQi
    nowSheet['C9'] =BaoGuanQiXian
    nowSheet['C10'] =MiJi

   #for row in range(1,11):
   #    for col in range(0,6):
   #        the_col = chr(65+col)
   #        cellNumber= the_col + str(row)
   #        print((nowSheet[cellNumber].value))

   #for row in range(1,11):
   #    for col in range(1,7):
   #        the_col = get_column_letter(col)
   #        cell_number =the_col+str(row)

    return  uniteBook

def copyDatatoTab1(mainTab:Workbook, subTab:Workbook, startIndex:str):
    mainsheet = mainTab['Sheet1']
    subsheet =subTab.active

    index = column_index_from_string(startIndex)
    B = get_column_letter(index +1)
    C = get_column_letter(index +2)

    mainsheet[B+'2'] = '档号：'
    mainsheet[B+'7'] ='立卷单位'
    mainsheet[B+'8'] ='起止日期'
    mainsheet[B+'9'] ='保管期限'
    mainsheet[B+'10'] ='密   级'

    mainsheet[C+'2'] = subsheet['C2'].value
    mainsheet[B+'6'] = "\n    "+subsheet['B6'].value     #首行缩进
    mainsheet[C+'7'] = subsheet['C7'].value
    mainsheet[C+'8'] = subsheet['C8'].value
    mainsheet[C+'9'] = subsheet['C9'].value
    mainsheet[C+'10']= subsheet['C10'].value

    #mainTab.save('test.xlsx')
    return mainTab

def decorateNewTab1(mainTab:Workbook, startIndex:str):
    mainsheet= mainTab['Sheet1']

    index =column_index_from_string(startIndex)

    A =get_column_letter(index)
    B =get_column_letter(index+1)
    C =get_column_letter(index+2)
    D =get_column_letter(index+3)
    E =get_column_letter(index+4)
    F =get_column_letter(index+5)

    columns =[A,B,C,D,E,F]

    #or row in range(1,11):
    #   for col in columns:
    #       cellIndex =col+str(row)
    #       mainsheet[cellIndex]

    #       pass

    # 设置列宽
    mainsheet.column_dimensions[A].width = 2.62962962962963
    mainsheet.column_dimensions[B].width = 15.7685185185185
    mainsheet.column_dimensions[C].width = 17.8888888888889
    mainsheet.column_dimensions[D].width = 17.8888888888889
    mainsheet.column_dimensions[E].width = 27.6296296296296
    mainsheet.column_dimensions[F].width = 2.62962962962963

    # 设置行高 磅
    mainsheet.row_dimensions[1].height = 35.25
    mainsheet.row_dimensions[2].height = 29.25
    mainsheet.row_dimensions[3].height = 29.25
    mainsheet.row_dimensions[4].height = 29.25
    mainsheet.row_dimensions[5].height = 49
    mainsheet.row_dimensions[6].height = 344
    mainsheet.row_dimensions[7].height = 42
    mainsheet.row_dimensions[8].height = 42
    mainsheet.row_dimensions[9].height = 42
    mainsheet.row_dimensions[10].height = 42


    font_18 =Font(
        name= '宋体',
        size= 18,
        bold=True
    )
    font_24 =Font(
        name= '宋体',
        size= 24,
        bold=True
    )
    font_16 =Font(
        name= '宋体',
        size= 16,
        bold=True
    )
    center =Alignment(
        horizontal='center',
        vertical='center'
    )
    left_center =Alignment(
        horizontal='left',
        vertical='center'
    )
    left_top =Alignment(
        wrapText=True,
        horizontal='left',
        vertical='top'
    )
    thin_border = Border(bottom=Side(style='thin'))

    mainsheet[B+'2'].font=font_18
    mainsheet[B+'2'].alignment =center

    mainsheet[C+'2'].font =font_18
    mainsheet[C+'2'].alignment =left_center

    mainsheet[B+'6'].font =font_24
    mainsheet[B+'6'].alignment =left_top


    for i in range(7,11):
        mainsheet[B + str(i)].font =font_18
        mainsheet[B + str(i)].alignment =center

        mainsheet[C + str(i)].font =font_16
        mainsheet[C + str(i)].alignment =center
        mainsheet[C + str(i)].border = thin_border


    # 合并单元格
    mainsheet.merge_cells(f"{B}1:{E}1")
    mainsheet.merge_cells(f"{C}2:{E}2")
    mainsheet.merge_cells(f"{B}3:{E}3")
    mainsheet.merge_cells(f"{B}5:{E}5")
    mainsheet.merge_cells(f"{B}6:{E}6")
    mainsheet.merge_cells(f"{C}7:{E}7")
    mainsheet.merge_cells(f"{C}8:{E}8")
    mainsheet.merge_cells(f"{C}9:{E}9")
    mainsheet.merge_cells(f"{C}10:{E}10")

    return  mainTab


# 写入到文件内,计划拆分，
# 先写入到一个WorKBook对象中，
# 然后再将WorkBook对象的数据复制到对应的位置，
# 最后再在文件的对应位置处更改单元格合并，并调整格式
def doWriteToTab1(data:dataTab1, filepath:str, start:str):

    mybook = load_workbook('test.xlsx')

    DangHao = data.DangHao
    AnJuanTiMing = data.AnJuanTiMing
    LiJuanDanWei = data.LiJuanDanWei
    QiZhiRiQi=data.QiZhiRiQi
    BaoGuanQiXian = data.BaoGuanQiXian
    MiJi = data.MiJi


    unitBook = oganizeDataTab1(DangHao, AnJuanTiMing, LiJuanDanWei, QiZhiRiQi, BaoGuanQiXian, MiJi)
    mybook = copyDatatoTab1(mybook, unitBook, start)
    mybook = decorateNewTab1(mainTab=mybook, startIndex=start)

    sheet = mybook["Sheet1"]
    sheet.page_margins = PageMargins(left=1.78 / 2.54,  # 转换为英寸
                                    right=1.78 / 2.54,
                                    top=1.91 / 2.54,
                                    bottom=1.91 / 2.54,
                                    header=0.76 / 2.54,
                                    footer=0.76 / 2.54)

    mybook.save(filepath)

    start = get_column_letter(column_index_from_string(start)+6)

    return start


if __name__ =="__main__":
    testData =dataTab1('danghao', 'test', 'Lijuandanwei', 'qizhiriqi', 'baoguanriqi', 'Miji')
    doWriteToTab1(data=testData, filepath='test.xlsx', start='A')


